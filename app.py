import io
import os
from datetime import date

import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="施工計画書 自動作成アプリ", layout="wide")

TEMPLATE_FILE = "01-1工事概要.xlsx"
TEMPLATE_SHEET = "1 -1.工事概要"

CELL_MAP = {
    "工事名": "C10",
    "工事番号": "C11",
    "工事場所": "C13",
    "受注者名": "C24",
}

st.title("施工計画書 自動作成アプリ")
st.write("入力フォームの内容を、指定のExcel様式に転記して出力します。")

with st.form("plan_form"):
    st.subheader("基本情報")
    project_name = st.text_input("工事名")
    project_number = st.text_input("工事番号")
    location = st.text_input("工事場所")
    contractor = st.text_input("受注者名")
    site_manager = st.text_input("現場代理人（現時点ではExcelに転記しません）")

    st.subheader("工期")
    start_date = st.date_input("工期（開始）", value=date.today())
    end_date = st.date_input("工期（終了）", value=date.today())

    st.subheader("施工計画")
    overview = st.text_area("工事概要", height=120)
    safety_plan = st.text_area("安全対策", height=120)
    quality_plan = st.text_area("品質管理", height=120)
    schedule_plan = st.text_area("工程計画", height=120)

    submitted = st.form_submit_button("Excelを作成する")

if submitted:
    input_data = {
        "工事名": project_name,
        "工事番号": project_number,
        "工事場所": location,
        "受注者名": contractor,
        "現場代理人": site_manager,
        "工期（開始）": str(start_date),
        "工期（終了）": str(end_date),
        "工事概要": overview,
        "安全対策": safety_plan,
        "品質管理": quality_plan,
        "工程計画": schedule_plan,
    }

    st.subheader("入力内容の確認")
    st.write(f"工事名: {input_data['工事名']}")
    st.write(f"工事番号: {input_data['工事番号']}")
    st.write(f"工事場所: {input_data['工事場所']}")
    st.write(f"受注者名: {input_data['受注者名']}")
    st.write(f"現場代理人: {input_data['現場代理人']}（今回は未転記）")

    if not os.path.exists(TEMPLATE_FILE):
        st.error(f"テンプレートファイルが見つかりません: {TEMPLATE_FILE}")
    else:
        try:
            wb = load_workbook(TEMPLATE_FILE)
        except Exception as e:
            st.error(f"テンプレートファイルを開けませんでした: {e}")
        else:
            if TEMPLATE_SHEET not in wb.sheetnames:
                st.error(f"シートが見つかりません: {TEMPLATE_SHEET}")
            else:
                ws = wb[TEMPLATE_SHEET]

                ws[CELL_MAP["工事名"]] = input_data["工事名"]
                ws[CELL_MAP["工事番号"]] = input_data["工事番号"]
                ws[CELL_MAP["工事場所"]] = input_data["工事場所"]
                ws[CELL_MAP["受注者名"]] = input_data["受注者名"]

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                st.success("Excelファイルを作成しました。")
                st.download_button(
                    label="転記済みExcelをダウンロード",
                    data=output,
                    file_name="01-1工事概要_転記済み.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
